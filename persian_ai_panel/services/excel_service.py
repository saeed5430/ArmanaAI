import pandas as pd
import openpyxl
import shutil
import os
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime


class ExcelService:
    def __init__(self):
        self.current_file: Optional[str] = None
        self.backup_dir = self._ensure_backup_dir()

    def _ensure_backup_dir(self) -> Path:
        for p in [Path("persian_ai_panel/backups/excel"), Path("backups/excel")]:
            p.mkdir(parents=True, exist_ok=True)
            return p
        Path("backups/excel").mkdir(parents=True, exist_ok=True)
        return Path("backups/excel")

    def import_excel(self, filepath: str) -> pd.DataFrame:
        try:
            df = pd.read_excel(filepath, engine='openpyxl')
            self.current_file = filepath
            self.create_backup()
            return df
        except Exception as e:
            raise Exception(f"Error reading Excel file: {e}")

    def validate_structure(self, df: pd.DataFrame) -> Dict[str, Any]:
        required_columns = ["نام محصول", "قیمت", "موجودی", "دسته‌بندی"]
        missing = [c for c in required_columns if c not in df.columns]
        duplicate_count = df.duplicated(subset=["نام محصول"]).sum() if "نام محصول" in df.columns else 0

        return {
            "is_valid": len(missing) == 0,
            "missing_columns": missing,
            "total_rows": len(df),
            "duplicate_count": duplicate_count,
            "columns": list(df.columns)
        }

    def create_backup(self):
        if not self.current_file or not os.path.exists(self.current_file):
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / f"backup_{ts}.xlsx"
        try:
            shutil.copy2(self.current_file, backup_path)
        except Exception as e:
            print(f"Backup error: {e}")

    def export_to_excel(self, data: List[Dict[str, Any]], filepath: str) -> bool:
        try:
            df = pd.DataFrame(data)
            df.to_excel(filepath, index=False, engine='openpyxl')
            return True
        except Exception as e:
            raise Exception(f"Error saving Excel: {e}")

    def get_excel_info(self, filepath: str) -> Dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            info = {
                "filename": Path(filepath).name,
                "sheets": wb.sheetnames,
                "active_sheet": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "file_size": os.path.getsize(filepath),
                "last_modified": datetime.fromtimestamp(os.path.getmtime(filepath))
            }
            wb.close()
            return info
        except Exception as e:
            raise Exception(f"Error reading file info: {e}")
